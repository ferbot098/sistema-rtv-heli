import streamlit as st
import pandas as pd
import os
import io
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import openpyxl
import hashlib

# --- CONFIGURACIÓN DE LA APP ---
st.set_page_config(page_title="RTV Helicópteros PRO", layout="wide", page_icon="🚁")
# --- SISTEMA DE SEGURIDAD Y ACCESO (EL GUARDIÁN) ---
# Aquí defines quién entra y cuál es su PIN. Luego puedes agregar a todos los pilotos.
# --- SISTEMA DE SEGURIDAD DESDE LA BÓVEDA ---
usuarios_pines = st.secrets["pilotos"]
usuarios = list(usuarios_pines.keys())

# Inicializamos el estado del candado
if 'autenticado' not in st.session_state:
    st.session_state.autenticado = False

# Si el candado está cerrado, mostramos la pantalla de acceso
if not st.session_state.autenticado:
    st.title("🔒 Acceso Restringido - Sistema RTV")
    st.markdown("Por favor, identifíquese para iniciar su turno operativo.")
    
    with st.form("formulario_login"):
        # 1. Aquí conectamos la caja desplegable con la nueva lista 'usuarios'
        usuario_intento = st.selectbox("Piloto / Operador", usuarios)
        pin_intento = st.text_input("PIN de 4 dígitos", type="password")
        boton_login = st.form_submit_button("Ingresar al Sistema", type="primary")
        
        if boton_login:
            # Encriptamos el intento del usuario en tiempo real
            pin_ingresado_hash = hashlib.sha256(pin_intento.encode()).hexdigest()
            
            # Comparamos hash contra hash
            if usuarios_pines.get(usuario_intento) == pin_ingresado_hash:
                st.session_state.autenticado = True
                st.session_state.usuario_actual = usuario_intento
                st.rerun() # Esto recarga la página y abre el candado
            else:
                st.error("❌ PIN incorrecto. Intente nuevamente.")
    
    # EL COMANDO LETAL: Si no está autenticado, la ejecución muere aquí.
    st.stop()

# ==========================================================
# SI PYTHON LLEGA A ESTA LÍNEA, EL USUARIO PASÓ EL FILTRO
# ==========================================================

# Botón para cerrar sesión (Lo ponemos en la barra lateral)
with st.sidebar:
    st.markdown(f"👤 **Usuario:** {st.session_state.usuario_actual}")
    if st.button("🚪 Cerrar Sesión"):
        st.session_state.autenticado = False
        st.rerun()
    st.divider()

# (AQUÍ CONTINÚA EL RESTO DE TU CÓDIGO INTACTO, DESDE "# --- CONTROL DE ARCHIVO ---")

# --- CONTROL DE ARCHIVO ---
archivo_excel = "RTV Y HEE PARA CELULAR-1-1.xlsx"

if not os.path.exists(archivo_excel):
    st.error(f"⛔ NO SE ENCUENTRA EL ARCHIVO: {archivo_excel}")
    st.stop()

# --- MOTOR DE DATOS (LECTURA INTELIGENTE) ---
@st.cache_data
def cargar_maestros():
    try:
        df = pd.read_excel(archivo_excel, sheet_name="DATOS", header=None, engine='openpyxl')
        pilotos = [p.strip() for p in df.iloc[3:, 1].dropna().astype(str).tolist() if p.lower() != 'nan' and len(p) > 2]
        aeronaves = [n.strip() for n in df.iloc[3:, 7].dropna().astype(str).tolist() if n.lower() != 'nan' and len(n) > 2]
        mab = [m.strip() for m in df.iloc[3:, 4].dropna().astype(str).tolist() if m.lower() != 'nan' and len(m) > 2]
        locaciones = [l.strip() for l in df.iloc[3:, 6].dropna().astype(str).tolist() if l.lower() != 'nan' and len(l) > 2]
        
        # GENERADOR MATEMÁTICO DE HORAS (00:00 a 23:59)
        lista_horas = [f"{h:02d}:{m:02d}" for h in range(24) for m in range(60)]
        
        return pilotos, aeronaves, mab, locaciones, lista_horas
    except Exception as e:
        st.error(f"Error crítico leyendo el Excel: {e}")
        return [], [], [], [], []

pilotos, aeronaves, mab, locaciones, lista_horas = cargar_maestros()

# --- FUNCIONES AUXILIARES ---
def calcular_minutos(hora_str):
    if pd.isna(hora_str) or ":" not in str(hora_str): return 0
    try:
        h, m = map(int, str(hora_str).split(':'))
        return h * 60 + m
    except: return 0

def minutos_a_hhmm(minutos):
    h = minutos // 60
    m = minutos % 60
    return f"{h:02d}:{m:02d}"

# --- FUNCIÓN ESCUDO: INYECTOR INTELIGENTE ---
def inyectar_valor(ws, celda, valor):
    try:
        ws[celda] = valor
    except AttributeError:
        # Si la celda es "read-only" (está combinada), busca la celda principal
        for rango in ws.merged_cells.ranges:
            if celda in rango:
                ws.cell(row=rango.min_row, column=rango.min_col).value = valor
                break

# --- MOTOR DE INYECCIÓN DE PLANTILLAS ---
def generar_excel_oficial(df_vuelos, nave, fecha, piloto, copiloto, mecanico):
    ruta_plantilla = "plantilla_rtv.xlsx"
    
    if not os.path.exists(ruta_plantilla):
        return None

    wb = openpyxl.load_workbook(ruta_plantilla)
    ws = wb["HOJA ESTADISTICA"]

    # 1. CABECERA (Usamos el inyector inteligente)
    inyectar_valor(ws, 'O8', "PLUSPETROL") 
    inyectar_valor(ws, 'V10', nave)
    inyectar_valor(ws, 'AQ10', fecha.strftime("%d/%m/%Y"))
    inyectar_valor(ws, 'P12', piloto)
    inyectar_valor(ws, 'AH12', copiloto)
    
    # 2. INYECTAR LA TABLA DE VUELOS
    fila_inicio = 27 
    
    for indice, fila_datos in df_vuelos.iterrows():
        fila_actual = fila_inicio + indice
        
        inyectar_valor(ws, f'AN{fila_actual}', fila_datos['Origen'])
        inyectar_valor(ws, f'BC{fila_actual}', fila_datos['Destino'])
        inyectar_valor(ws, f'AI{fila_actual}', fila_datos['H. Salida'])
        inyectar_valor(ws, f'AX{fila_actual}', fila_datos['H. Llegada'])
        inyectar_valor(ws, f'V{fila_actual}', fila_datos['Pax'])
        inyectar_valor(ws, f'Y{fila_actual}', fila_datos['Carga (Kg)'])
        inyectar_valor(ws, f'F{fila_actual}', fila_datos['Comb. Salida (Gal)'])
        inyectar_valor(ws, f'R{fila_actual}', fila_datos['Comb. Llegada (Gal)'])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
# --- BARRA LATERAL ---
with st.sidebar:
    st.header("📋 Datos del Día")
    nave_selec = st.selectbox("Aeronave", aeronaves)
    fecha_selec = st.date_input("Fecha de Vuelo", datetime.now())
    piloto_selec = st.selectbox("Piloto al Mando (PIC)", pilotos)
    copiloto_selec = st.selectbox("Copiloto (SIC)", ["-"] + pilotos)
    mecanico_selec = st.selectbox("Técnico / MAB", mab)

# --- ESTADO DE SESIÓN Y TABLA (AHORA CON HORAS DESPLEGABLES) ---
if 'bitacora' not in st.session_state:
    st.session_state.bitacora = pd.DataFrame(
        [{"Origen": locaciones[0] if locaciones else "-", "Destino": locaciones[0] if locaciones else "-", 
          "H. Salida": "08:00", "H. Llegada": "09:00", "Comb. Salida (Gal)": 0, "Comb. Llegada (Gal)": 0, "Pax": 0, "Carga (Kg)": 0}]
    )
if 'calculo_exitoso' not in st.session_state:
    st.session_state.calculo_exitoso = False
if 'df_exportable' not in st.session_state:
    st.session_state.df_exportable = None

st.subheader("Registro de Tramos")

# TABLA INTELIGENTE CON SELECTORES ABSOLUTOS
df_editado = st.data_editor(
    st.session_state.bitacora, 
    num_rows="dynamic", 
    use_container_width=True,
    column_config={
        "Origen": st.column_config.SelectboxColumn("Origen", options=locaciones, required=True),
        "Destino": st.column_config.SelectboxColumn("Destino", options=locaciones, required=True),
        "H. Salida": st.column_config.SelectboxColumn("H. Salida", options=lista_horas, required=True),
        "H. Llegada": st.column_config.SelectboxColumn("H. Llegada", options=lista_horas, required=True),
        "Comb. Salida (Gal)": st.column_config.NumberColumn("Comb. Inicial", min_value=0),
        "Comb. Llegada (Gal)": st.column_config.NumberColumn("Comb. Final", min_value=0),
    }
)

st.divider()

# --- BOTÓN DE PROCESAMIENTO ---
if st.button("CALCULAR TOTALES Y VALIDAR", type="primary"):
    total_minutos = total_consumo = total_pax = total_carga = 0
    errores = []

    for i, row in df_editado.iterrows():
        m_salida = calcular_minutos(row["H. Salida"])
        m_llegada = calcular_minutos(row["H. Llegada"])
        
        if m_salida > 0 and m_llegada > 0:
            if m_llegada < m_salida:
                errores.append(f"Fila {i+1}: La hora de llegada es anterior a la salida.")
            else:
                total_minutos += (m_llegada - m_salida)
        
        c_salida = row["Comb. Salida (Gal)"]
        c_llegada = row["Comb. Llegada (Gal)"]
        consumo = c_salida - c_llegada
        
        if consumo < 0:
            errores.append(f"Fila {i+1}: ¡Generaste combustible! Verifica datos.")
        else:
            total_consumo += consumo

        total_pax += row["Pax"]
        total_carga += row["Carga (Kg)"]

    if errores:
        for e in errores: st.error(e)
        st.session_state.calculo_exitoso = False
    else:
        st.markdown(f"### 📊 Resumen de Operación ({len(df_editado)} tramos)")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("⏱️ Tiempo Total", minutos_a_hhmm(total_minutos))
        c2.metric("⛽ Consumo Total", f"{total_consumo} Gal")
        c3.metric("👥 Pasajeros", f"{total_pax}")
        c4.metric("📦 Carga Total", f"{total_carga} Kg")
        
        st.success("✅ Registro validado correctamente.")
        
        df_export = df_editado.copy()
        df_export.insert(0, "Fecha", fecha_selec)
        df_export.insert(1, "Aeronave", nave_selec)
        df_export.insert(2, "Piloto", piloto_selec)
        
        st.session_state.df_exportable = df_export
        st.session_state.calculo_exitoso = True

# --- ZONA DE EXPORTACIÓN ---
if st.session_state.calculo_exitoso and st.session_state.df_exportable is not None:
    st.divider()
    st.markdown("### 📥 Opciones de Exportación")
    
    # 1. LLAMAMOS AL NUEVO MOTOR DE PLANTILLAS
    excel_listo = generar_excel_oficial(
        df_vuelos = st.session_state.df_exportable,
        nave = nave_selec,
        fecha = fecha_selec,
        piloto = piloto_selec,
        copiloto = copiloto_selec,
        mecanico = mecanico_selec
    )
    
    col_a, col_b = st.columns(2)
    
    # 2. BOTÓN IZQUIERDO: DESCARGA EL FORMATO OFICIAL (Si la plantilla existe)
    with col_a:
        if excel_listo:
            nombre_arch = f"HOJA_ESTADISTICA_{nave_selec}_{fecha_selec}.xlsx"
            st.download_button(
                label="📄 DESCARGAR FORMATO OFICIAL",
                data=excel_listo,
                file_name=nombre_arch,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
        else:
            st.warning("⚠️ Crea 'plantilla_rtv.xlsx' para habilitar la descarga.")
            
    # 3. BOTÓN DERECHO: GUARDA EN EL HISTORIAL (Intacto)
    with col_b:
        if st.button("💾 GUARDAR EN BASE DE DATOS HISTÓRICA"):
            archivo_historial = "BASE_DE_DATOS_VUELOS.csv"
            if not os.path.exists(archivo_historial):
                st.session_state.df_exportable.to_csv(archivo_historial, index=False)
            else:
                st.session_state.df_exportable.to_csv(archivo_historial, mode='a', header=False, index=False)
            st.toast("✅ Vuelo guardado en el historial de la empresa.")

# --- HISTORIAL GENERAL (Intacto) ---
st.divider()
if st.checkbox("Ver Base de Datos Histórica"):
    if os.path.exists("BASE_DE_DATOS_VUELOS.csv"):
        st.dataframe(pd.read_csv("BASE_DE_DATOS_VUELOS.csv"))
    else:
        st.info("Aún no hay registros en la base de datos.")